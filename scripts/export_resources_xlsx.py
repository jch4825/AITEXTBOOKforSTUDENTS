from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CATEGORY_ORDER = ['policy', 'ai-basics', 'ethics', 'lesson', 'assessment', 'research-etc']
CATEGORY_DISPLAY_TITLE = {
    'policy': '정책·지침',
    'ai-basics': 'AI 기초 이해',
    'ethics': 'AI 윤리·안전',
    'lesson': '수업 지원',
    'assessment': '행정 업무',
    'research-etc': '연구·연수·기타',
}

categories = [
    ('ai-basics', [
        ('1. 대화형 AI', [
            ('ChatGPT', 'https://chatgpt.com', 'OpenAI의 대화형 AI. GPT-4o 기반의 범용 AI 어시스턴트'),
            ('Gemini', 'https://gemini.google.com', 'Google의 대화형 AI. 구글 서비스·드라이브 연동에 강점'),
            ('Claude', 'https://claude.ai', 'Anthropic의 대화형 AI. 긴 문서 분석과 안전성·정확성에 강점'),
            ('Grok', 'https://grok.com', 'xAI(일론 머스크)의 AI. X(트위터) 실시간 데이터 연동'),
            ('Perplexity', 'https://www.perplexity.ai', '실시간 웹 검색 기반 AI 검색 엔진. 출처를 함께 제공'),
            ('젠스파크 (Genspark)', 'https://www.genspark.ai', '멀티 AI 에이전트 기반 검색·정리 도구. 복잡한 주제를 자동 요약'),
        ]),
        ('2. Google AI 도구', [
            ('Google AI Studio', 'https://aistudio.google.com', 'Gemini API를 무료로 테스트하고 키를 발급받는 구글 공식 개발자 도구'),
            ('Google Colab', 'https://colab.research.google.com', '브라우저에서 Python 코드를 실행하는 무료 클라우드 노트북. GPU 무료 제공'),
            ('Google Vids', 'https://workspace.google.com/products/vids/', 'Google Workspace의 AI 영상 제작 도구. 텍스트로 영상 슬라이드 자동 생성'),
            ('Google Flow', 'https://flow.google.com', 'Google의 AI 영화 제작 도구. Veo 2 기반으로 고품질 영상 생성'),
            ('Antigravity', 'https://antigravity.google/', 'Google의 전문가용 바이브코딩 도구'),
            ('NotebookLM', 'https://notebooklm.google.com', '문서·PDF를 업로드하면 AI가 요약·질의응답해 주는 구글의 AI 노트 도구'),
        ]),
        ('3. 나만의 로컬 인공지능 스튜디오', [
            ('Ollama', 'https://ollama.com', '인터넷 없이 로컬 PC에서 오픈소스 LLM을 명령어로 실행하는 도구'),
            ('LM Studio', 'https://lmstudio.ai', 'GUI로 로컬 AI 모델을 쉽게 내려받고 실행. 비개발자도 사용 가능'),
            ('GPT4ALL', 'https://gpt4all.io', '저사양 PC에서도 실행 가능한 로컬 AI 플랫폼. 다양한 모델 지원'),
            ('Jan', 'https://jan.ai', '오픈소스 로컬 AI 클라이언트. ChatGPT와 유사한 UI로 로컬 모델 실행'),
            ('Open WebUI', 'https://openwebui.com', 'Ollama 등 로컬 모델에 연결하는 웹 기반 인터페이스. 다중 모델 관리 가능'),
        ]),
        ('4. 음성 및 소리 제작', [
            ('ElevenLabs', 'https://elevenlabs.io', '텍스트를 자연스러운 목소리로 변환하는 AI TTS 서비스. 다국어 지원'),
            ('클로바 더빙', 'https://clovadubbing.naver.com', '네이버의 한국어 특화 AI 더빙·TTS 서비스. 영상에 자동 더빙 적용 가능'),
            ('Suno AI', 'https://suno.com', '텍스트 프롬프트만으로 완성된 노래(가사+반주)를 생성하는 AI 작곡 도구'),
            ('Udio', 'https://www.udio.com', '고품질 AI 음악 생성 도구. 다양한 장르와 세밀한 스타일 조정 지원'),
        ]),
        ('5. No Code AI 도구', [
            ('Teachable Machine', 'https://teachablemachine.withgoogle.com', 'Google의 코딩 없는 머신러닝 체험 도구. 이미지·소리·자세를 학습시켜 AI 모델 직접 제작'),
            ('orange3', 'https://orangedatamining.com', '드래그앤드롭 방식의 오픈소스 데이터 분석·머신러닝 도구. 시각적 워크플로우로 AI 학습 가능'),
            ('lobe', 'https://www.lobe.ai', 'Microsoft의 코드 없는 이미지 분류 AI 학습 도구. 사진만 모으면 분류 모델 생성 가능'),
            ('Brightics AI (삼성)', 'https://www.brightics.ai', '삼성SDS의 노코드 AI 분석 플랫폼. 제조·유통·금융 데이터 분석에 특화된 기업용 도구'),
        ]),
        ('6. 제작·디자인 도구', [
            ('Gamma', 'https://gamma.app', 'AI로 프레젠테이션·문서·웹페이지를 빠르게 제작. 텍스트 입력만으로 슬라이드 자동 생성'),
            ('Canva', 'https://www.canva.com', '드래그앤드롭 방식의 범용 디자인 도구. Magic Write 등 AI 기능 내장'),
            ('미리캔버스', 'https://www.miricanvas.com', '한국어 특화 무료 디자인 도구. 교육용 템플릿 다수 제공'),
            ('망고보드', 'https://www.mangoboard.net', '한국형 인포그래픽·카드뉴스·PPT 제작 도구. 교육 현장에서 많이 활용'),
            ('투닝', 'https://tooning.io', 'AI 기반 웹툰·만화 제작 도구. 캐릭터와 배경을 자동 생성해 학습 만화 제작 가능'),
            ('Figma', 'https://www.figma.com', 'UI/UX 디자인 협업 도구. AI 플러그인 다수 지원. 교육용 무료 플랜 제공'),
        ]),
    ]),
    ('research-etc', [
        ('1. 연구', [
            ('AIEDAP AI융합교육 학술자료', 'https://aiedap.or.kr/ai융합수업-지원/ai융합교육-학술자료/', 'AI 융합교육 관련 학술 논문·연구 보고서를 모아둔 AIEDAP 공식 페이지'),
            ('한국과학창의재단 연구보고서', 'https://www.kosac.re.kr/menus/244/boards/457/posts', '한국과학창의재단이 발간한 과학·수학·정보·융합 교육 분야 연구보고서 모음'),
            ('한국인공지능교육학회', 'http://aied.or.kr/', '인공지능 교육 연구를 전문으로 하는 학술 단체. 학술지·학술대회 정보 제공'),
        ]),
        ('2. 연수', [
            ('KERIS 종합교육연수원 연수 자료집', 'https://www.cet.keris.or.kr/usr/reference/selectTrainingListPageVw.do?menuNo=23100', 'KERIS 종합교육연수원에서 제공하는 교원 연수 자료집 모음'),
        ]),
        ('3. 기타', [
            ('교과 학습발달상황 작성 GPT', 'https://chatgpt.com/g/g-6753f78947708191b221bbf659619dfd-seongcwigijun-giban-gyogwahagseubbaldalsanghwang-saengseong-caesbos-jeongwamog', 'ChatGPT 커스텀 GPT. 학생의 학습 활동·성과를 입력하면 교과 학습발달상황을 자동으로 생성'),
            ('교육 아이디어 도우미 GPT', 'https://chatgpt.com/g/g-677cab3951bc8191adf633bd9e6b9fcb-gyoyug-aidieo-doumi', 'ChatGPT 커스텀 GPT. 수업 주제·상황을 입력하면 교육 활동 아이디어를 제안해 주는 도우미'),
            ('교육·다큐멘터리·어학·어린이 유튜브 영상 보기 GPT', 'https://chatgpt.com/g/g-f7aSrkPEP-gyoyug-dakyumenteori-eohag-eorini-yutyubeu-yeongsang-bogi-education', 'ChatGPT 커스텀 GPT. 교육·다큐멘터리·어학·어린이 분야의 유튜브 영상을 추천하고 함께 시청·정리할 수 있는 도우미'),
        ]),
    ]),
    ('lesson', [
        ('1. 한글·문해 수업 자료', [
            ('온한글', 'https://onhan.cne.go.kr/', '충청남도교육청 한글 해득 진단·보정 시스템. 미해득 학생 보정 자료와 성장 기록 지원'),
        ]),
        ('2. 창작·스몰 도구', [
            ('Chrome Music Lab', 'https://musiclab.chromeexperiments.com/', '리듬, 멜로디, 소리의 파형을 놀이처럼 탐색하는 구글의 웹 기반 음악 실험실'),
            ('AutoDraw', 'https://www.autodraw.com/', '간단한 스케치를 AI가 그림 아이콘으로 추천해 주는 구글 크리에이티브 랩의 드로잉 도구'),
        ]),
        ('3. 프로그래밍 수업 도구', [
            ('Delightex Edu', 'https://www.delightex.com/edu', '3D 공간, VR·AR 콘텐츠를 만들고 블록 코딩과 스크립트로 상호작용을 구현하는 교육용 제작 도구'),
            ('엔트리', 'https://playentry.org/ko/', '네이버 커넥트재단에서 운영하는 비영리 블록 코딩 교육 플랫폼'),
            ('스크래치', 'https://scratch.mit.edu/', '스토리, 게임, 애니메이션을 만들며 프로그래밍 기초를 배우는 블록 기반 창작 플랫폼'),
        ]),
        ('4. 시뮬레이션 수업 도구', [
            ('자바실험실', 'https://javalab.org/ko/', '과학·수학 개념을 웹 시뮬레이션으로 탐구할 수 있는 한국어 실험 자료 사이트'),
            ('지오지브라', 'https://www.geogebra.org/?lang=ko', '함수, 기하, 통계, 3D 그래프 등을 조작하며 탐구하는 수학 시뮬레이션 플랫폼'),
            ('알지오매스', 'https://www.algeomath.kr/', '수학 개념을 그래프와 기하 모델로 시각화해 수업에 활용할 수 있는 국내 수학 탐구 도구'),
            ('VlabON', 'https://www.vlabon.re.kr/', '가상 실험과 과학 탐구 활동을 지원하는 온라인 과학 실험 플랫폼'),
        ]),
        ('5. AI 교육 포털', [
            ('에듀넷 AI·SW 교육', 'https://ai.edunet.net', '교육부 공식 AI·SW 교육 포털. 교사·학생용 자료, 연수, 수업 지원 통합 제공'),
            ('에듀넷 티-클리어 AI 교수학습자료', 'https://ai.edunet.net/aiTchLrngData/list/408', '에듀넷의 AI 교수·학습 자료 목록. 수업에서 바로 활용 가능한 자료 모음'),
            ('에듀넷 AI·SW 연구·기타자료', 'https://ai.edunet.net/aiStdyEtcData/list/411', 'AI·SW 교육 관련 연구 보고서 및 기타 참고 자료'),
            ('한국과학창의재단 SAI', 'https://www.kosac.re.kr/menus/1124/contents/1124', '한국과학창의재단의 학교 AI 교육(SAI) 사업 안내 및 자료 제공 페이지'),
            ('SAI 교육콘텐츠', 'https://sai.software.kr', '학교 AI 교육을 위한 교육콘텐츠·수업 자료 포털'),
            ('AI·디지털 교육자료 수업지원센터', 'https://www.ai-dt.net', '교사가 AI·디지털 교육자료를 찾고 수업에 바로 적용할 수 있는 지원 센터'),
            ('AI·디지털 교육자료(aidtbook)', 'https://www.aidtbook.kr/index.do', 'AI 디지털 교과서 관련 교육자료 통합 제공 포털'),
            ('미래엔 AI·디지털 교육자료', 'https://aidt.m-teacher.co.kr', '출판사 미래엔에서 제공하는 AI·디지털 교육자료 및 교사 지원 자료'),
            ('아이스크림 AI 디지털 교육자료', 'https://aidt.i-scream.co.kr', '아이스크림미디어의 AI 디지털 교육자료 플랫폼. 초등 특화 콘텐츠 다수'),
            ('빛고을 아이(AI) — 광주 AI 교육 프로그램', 'https://sw.gen.go.kr/bbs/board.php?bo_table=data&wr_id=15', '광주 AI 교육센터가 2024년에 공개한 초등용 AI 교육 프로그램 교재'),
            ('울산교육청 우리아이', 'https://wooriai.use.go.kr/', '울산광역시교육청에서 운영하는 학생용 AI 챗봇 서비스'),
            ('초·중등 AI 수학 융합교육 교수학습자료', 'https://sw.gen.go.kr/bbs/board.php?bo_table=data&wr_id=64', '광주 SW 교육지원센터가 2024년에 발간한 초·중등 인공지능-수학 융합 수업 자료'),
            ('초등인공지능교육.COM', 'https://www.xn--ob0bug89lftci99a2oa25k5pi.com/', '초등 교사가 직접 운영하는 인공지능 수업 자료 모음 사이트'),
            ('VIVASAM 초등 인공지능 자료', 'https://e.vivasam.com/creative/edu/sw/cia/list', '비상교육 비바샘에서 제공하는 초등 인공지능·창의 수업 자료 모음'),
            ('초등 인공지능 교육 모듈형 39강좌', 'https://sites.google.com/ptsaebit.es.kr/metaverse/%EC%B4%88%EB%93%B1swai-%EC%BD%98%ED%85%90%EC%B8%A0-%EB%AA%A8%EC%9D%8C/%EC%B4%88%EB%93%B1-%EC%9D%B8%EA%B3%B5%EC%A7%80%EB%8A%A5-%EA%B5%90%EC%9C%A1%EB%AA%A8%EB%93%88%ED%98%95-39%EA%B0%95%EC%A2%8C', '초등 단계의 인공지능 교육을 모듈형 39강좌로 구성한 교사용 콘텐츠 모음'),
            ('소프트웨어야 놀자', 'https://www.playsw.or.kr/main', '네이버커넥트재단이 운영하는 SW·AI 교육 콘텐츠 플랫폼'),
            ('씨마스 AI 디지털 교육자료', 'https://viewer.cmass.kr/html/aidt/aidt_e/view_aidt_e.html', '씨마스가 제공하는 AI 디지털 교과서 미리보기·교육자료'),
            ('천재교육 AI 디지털 교육자료 지원센터', 'https://support.aitextbook.co.kr/', '천재교육의 AI 디지털 교과서 활용 안내 및 교사 지원 자료'),
            ('MyAI Edu', 'https://myai.kr/', '교사·학생을 위한 인공지능 교육 통합 콘텐츠 사이트'),
            ('인공지능 교육자료 사이트 모음(노션)', 'https://ai4edu.notion.site/AI-92b8c8eca0294fc7a9a2538bca205bdf', '이준기·김귀훈 선생님이 정리한 국내외 AI 교육자료 링크 모음 노션 페이지'),
        ]),
    ]),
    ('assessment', [
        ('학교업무 도움자료', [
            ('경상남도교육청 학교업무 도움자료', 'https://hryoon0.github.io/helppage/', '경상남도교육청 학교업무 도움자료 사이트. 초등학교 업무 참고 자료를 확인할 수 있습니다.'),
        ]),
    ]),
    ('ethics', []),
    ('policy', [
        ('1. 국가 및 교육부 지침', [
            ('인공지능 행동계획', 'https://drive.google.com/file/d/1uVEhE63fvCxCf3eukHIHmMi7vZUvu1jS/view?usp=drive_link', '교육부의 인공지능 교육 추진 방향과 단계별 실행 계획을 담은 공식 문서'),
            ('초중등 디지털 인프라 관리 가이드', 'https://www.moe.go.kr/boardCnts/viewRenew.do?boardID=316&boardSeq=104618&lev=0&searchType=null&statusYN=W&page=3&s=moe&m=0302&opType=N', '교육부가 발표한 초·중등학교 디지털 인프라 구축·운영·관리 표준 가이드'),
            ('제7차 교육정보화 기본계획', 'https://www.moe.go.kr/boardCnts/viewRenew.do?boardID=351&boardSeq=103106&lev=0&searchType=null&statusYN=W&page=1&s=moe&m=0310&opType=N', '교육부가 수립한 제7차 교육정보화 기본계획. 교육정보화의 비전과 중점 추진 과제를 담은 공식 정책 문서'),
            ('2025년 지능정보사회 실행계획', 'https://www.moe.go.kr/boardCnts/viewRenew.do?boardID=351&boardSeq=103108&lev=0&searchType=null&statusYN=W&page=1&s=moe&m=0310&opType=N', '교육부의 2025년 지능정보사회 실행계획. 교육 분야 AI·디지털 전환을 위한 연도별 추진 과제'),
        ]),
        ('2. 시도교육청 자료', [
            ('서울시교육청 AIEDAP AI 융합수업 우수 사례집', 'https://www.sen.go.kr/user/bbs/BD_selectBbs.do?q_bbsSn=1462&q_bbsDocNo=20240601181736054', '서울시교육청 AIEDAP 사업의 AI 융합수업 우수 실천 사례를 모은 자료집'),
            ('생성형 인공지능(AI) 개발·활용을 위한 개인정보 처리 안내서(2025.8.)', 'https://www.gne.go.kr/user/bbs/BD_selectBbs.do?q_rowPerPage=10&q_currPage=1&q_sortName=&q_sortOrder=&q_bbsSn=1464&q_bbsDocNo=20260424101701193&q_menu=&q_ctgryCd=&q_clsfNo=1&q_lwrkCdId=10&q_searchKeyTy=ttl___1002&q_searchVal=&', '생성형 AI 개발·활용 단계에서 지켜야 할 개인정보 처리 원칙과 실무 절차를 정리한 2025년 8월 안내서'),
            ('경상남도교육청 AI·디지털 교육 종합 추진 계획', 'https://www.gne.go.kr/user/bbs/BD_selectBbs.do?q_rowPerPage=10&q_currPage=2&q_sortName=&q_sortOrder=&q_bbsSn=1464&q_bbsDocNo=20260418223116303&q_menu=&q_ctgryCd=&q_clsfNo=1&q_lwrkCdId=10&q_searchKeyTy=ttl___1002&q_searchVal=&', '경상남도교육청이 발표한 AI·디지털 교육의 비전과 단계별 실행 과제를 담은 종합 추진 계획'),
            ('2026년 AI 중점학교 기본계획 및 AI 중점학교 명단', 'https://www.gne.go.kr/user/bbs/BD_selectBbs.do?q_rowPerPage=10&q_currPage=2&q_sortName=&q_sortOrder=&q_bbsSn=1464&q_bbsDocNo=20260309200422066&q_menu=&q_ctgryCd=&q_clsfNo=1&q_lwrkCdId=10&q_searchKeyTy=ttl___1002&q_searchVal=&', '경상남도교육청 2026년 AI 중점학교 운영 기본계획과 선정 학교 명단'),
            ('2025년 정보통신윤리교육 강화 및 지능정보서비스 과의존 예방교육 기본계획', 'https://www.gne.go.kr/user/bbs/BD_selectBbs.do?q_rowPerPage=10&q_currPage=10&q_sortName=&q_sortOrder=&q_bbsSn=1464&q_bbsDocNo=20250326085906790&q_menu=&q_ctgryCd=&q_clsfNo=1&q_lwrkCdId=10&q_searchKeyTy=ttl___1002&q_searchVal=&', '경상남도교육청 2025년 정보통신윤리교육 강화와 지능정보서비스 과의존 예방교육 추진 기본계획'),
        ]),
    ]),
]

cat_map = dict(categories)
ordered = [(cid, CATEGORY_DISPLAY_TITLE[cid], cat_map[cid]) for cid in CATEGORY_ORDER]

wb = Workbook()
ws = wb.active
ws.title = '자료실'

headers = ['번호', '대분류', '하위 분류', '자료명', 'URL', '설명']
ws.append(headers)

header_font = Font(name='맑은 고딕', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', start_color='4F46E5')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin = Side(border_style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

body_font = Font(name='맑은 고딕', size=10)
url_font = Font(name='맑은 고딕', size=10, color='1155CC', underline='single')
wrap_align = Alignment(vertical='top', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='top', wrap_text=True)

row_no = 2
seq = 1
for _cid, cat_title, subs in ordered:
    for sub_label, items in subs:
        for title, url, desc in items:
            ws.cell(row=row_no, column=1, value=seq).alignment = center_align
            ws.cell(row=row_no, column=2, value=cat_title).alignment = wrap_align
            ws.cell(row=row_no, column=3, value=sub_label).alignment = wrap_align
            ws.cell(row=row_no, column=4, value=title).alignment = wrap_align
            url_cell = ws.cell(row=row_no, column=5, value=url)
            url_cell.hyperlink = url
            url_cell.font = url_font
            url_cell.alignment = wrap_align
            ws.cell(row=row_no, column=6, value=desc).alignment = wrap_align
            for c in range(1, 7):
                cur = ws.cell(row=row_no, column=c)
                cur.border = border
                if c != 5:
                    cur.font = body_font
            seq += 1
            row_no += 1

widths = [6, 14, 26, 38, 60, 60]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 28
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:F{row_no - 1}'

out = r'C:\AI\AI_bridge_test_v0.1\AI_Bridge_자료실_링크.xlsx'
wb.save(out)
print(f'saved: {out}, rows: {row_no - 1}')
