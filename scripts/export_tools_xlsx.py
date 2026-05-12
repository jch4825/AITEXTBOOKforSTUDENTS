import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "AI_Bridge_AI_도구_목록.json"
OUT = ROOT / "AI_Bridge_AI_도구_목록.xlsx"


def join_inputs(inputs):
    if not inputs:
        return ""

    lines = []
    for item in inputs:
        required = "필수" if item["required"] else "선택"
        line = f'[{required}] {item["label"]} ({item["type"]})'
        details = []
        if item["placeholder"]:
            details.append(f'placeholder: {item["placeholder"]}')
        if item["hint"]:
            details.append(f'hint: {item["hint"]}')
        if item["options"]:
            details.append(f'options:\n{item["options"]}')
        if details:
            line += "\n  " + "\n  ".join(details)
        lines.append(line)
    return "\n\n".join(lines)


with SOURCE.open("r", encoding="utf-8") as f:
    tools = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "AI 도구 목록"

headers = [
    "번호",
    "도구 ID",
    "분류",
    "세부 분류",
    "도구 명칭",
    "설명",
    "도구 유형",
    "외부 URL",
    "호스트",
    "태그",
    "연계 레슨",
    "필수 입력값",
    "전체 입력값",
    "구현 프롬프트",
]
ws.append(headers)

for tool in tools:
    ws.append([
        tool["no"],
        tool["id"],
        tool["category"],
        tool["subCategory"],
        tool["title"],
        tool["description"],
        tool["kind"],
        tool["externalUrl"],
        tool["hostLabel"],
        tool["tags"],
        tool["usedInLessons"],
        tool["requiredInputs"],
        join_inputs(tool["inputs"]),
        tool["implementationPrompt"],
    ])

header_fill = PatternFill("solid", fgColor="111827")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border

widths = {
    "A": 8,
    "B": 26,
    "C": 18,
    "D": 28,
    "E": 34,
    "F": 48,
    "G": 12,
    "H": 52,
    "I": 18,
    "J": 34,
    "K": 18,
    "L": 34,
    "M": 72,
    "N": 100,
}

for col, width in widths.items():
    ws.column_dimensions[col].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

for row_idx in range(2, ws.max_row + 1):
    ws.row_dimensions[row_idx].height = 90

summary = wb.create_sheet("요약")
summary.append(["항목", "값"])
summary.append(["총 도구 수", len(tools)])
summary.append(["내부 실행형 도구 수", sum(1 for tool in tools if tool["kind"] == "api")])
summary.append(["외부 링크 도구 수", sum(1 for tool in tools if tool["kind"] == "external")])
summary.append(["생성 파일", OUT.name])
summary.append(["원본", "src/tools/ToolRegistry.ts"])

for cell in summary[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

for row in summary.iter_rows(min_row=2):
    for cell in row:
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

summary.column_dimensions["A"].width = 24
summary.column_dimensions["B"].width = 48

wb.save(OUT)
print(OUT)
